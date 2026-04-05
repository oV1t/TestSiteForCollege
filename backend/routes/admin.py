from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select, func, SQLModel
from sqlalchemy.exc import IntegrityError
from typing import List
import io
import csv
from datetime import datetime
from models import Choice, ChoiceSet, User, Discipline, UserRole, Campaign
from database import get_session
from routes.auth import require_admin

router = APIRouter()

@router.get("/stats")
def get_stats(
    session: Session = Depends(get_session),
    admin: User = Depends(require_admin)
):
    # Total students who made a choice
    total_participants = session.exec(select(func.count(ChoiceSet.id))).one()
    
    # Discipline popularity
    disciplines = session.exec(select(Discipline)).all()
    stats = []
    for d in disciplines:
        count_p1 = session.exec(select(func.count(Choice.id)).where(Choice.discipline_id == d.id, Choice.priority == 1)).one()
        count_p2 = session.exec(select(func.count(Choice.id)).where(Choice.discipline_id == d.id, Choice.priority == 2)).one()
        count_p3 = session.exec(select(func.count(Choice.id)).where(Choice.discipline_id == d.id, Choice.priority == 3)).one()
        
        # New: Group breakdown (Total for all priorities)
        group_counts = session.exec(
            select(User.group_name, func.count(Choice.id))
            .join(ChoiceSet, ChoiceSet.user_id == User.id)
            .join(Choice, Choice.choice_set_id == ChoiceSet.id)
            .where(Choice.discipline_id == d.id)
            .group_by(User.group_name)
        ).all()
        
        group_stats = [{"group": g, "count": c} for g, c in group_counts if g]
        stats.append({
            "id": d.id,
            "code": d.code,
            "title": d.title,
            "priority1": count_p1,
            "priority2": count_p2,
            "priority3": count_p3,
            "total": count_p1 + count_p2 + count_p3,
            "group_stats": group_stats
        })
    
    return {
        "total_participants": total_participants,
        "discipline_stats": stats
    }

@router.get("/export/csv")
def export_choices_csv(
    session: Session = Depends(get_session),
    admin: User = Depends(require_admin)
):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Email", "Full Name", "Group", "Priority 1", "Priority 2", "Priority 3", "Submitted At"])
    
    choicesets = session.exec(select(ChoiceSet)).all()
    for cs in choicesets:
        p1 = next((c.discipline.title for c in cs.choices if c.priority == 1), "")
        p2 = next((c.discipline.title for c in cs.choices if c.priority == 2), "")
        p3 = next((c.discipline.title for c in cs.choices if c.priority == 3), "")
        writer.writerow([
            cs.user.email,
            cs.user.full_name,
            cs.user.group_name,
            p1, p2, p3,
            cs.submitted_at.isoformat()
        ])
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=elective_choices.csv"}
    )

from pydantic import BaseModel
from typing import Optional

class DisciplineCreate(BaseModel):
    code: str
    title: str
    short_info: Optional[str] = None
    doc_url: Optional[str] = None
    commission_name: Optional[str] = None
    specialty_code: Optional[str] = None
    credits: Optional[float] = None
    teacher_name: Optional[str] = None
    competence_type: Optional[str] = None
    active: bool = True

class DisciplineUpdate(BaseModel):
    code: Optional[str] = None
    title: Optional[str] = None
    short_info: Optional[str] = None
    doc_url: Optional[str] = None
    commission_name: Optional[str] = None
    specialty_code: Optional[str] = None
    credits: Optional[float] = None
    teacher_name: Optional[str] = None
    competence_type: Optional[str] = None
    active: Optional[bool] = None

@router.get("/disciplines")
async def get_admin_disciplines(session: Session = Depends(get_session), admin: User = Depends(require_admin)):
    return session.exec(select(Discipline).order_by(func.length(Discipline.code), Discipline.code)).all()

@router.post("/disciplines", response_model=Discipline)
def create_discipline(
    data: DisciplineCreate,
    session: Session = Depends(get_session),
    admin: User = Depends(require_admin)
):
    new_discipline = Discipline(**data.model_dump())
    session.add(new_discipline)
    session.commit()
    session.refresh(new_discipline)
    return new_discipline

@router.put("/disciplines/{discipline_id}", response_model=Discipline)
def update_discipline(
    discipline_id: int,
    data: DisciplineUpdate,
    session: Session = Depends(get_session),
    admin: User = Depends(require_admin)
):
    discipline = session.get(Discipline, discipline_id)
    if not discipline:
        raise HTTPException(status_code=404, detail="Discipline not found")
    
    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(discipline, key, value)
    
    discipline.updated_at = datetime.utcnow()
    session.add(discipline)
    session.commit()
    session.refresh(discipline)
    return discipline

@router.delete("/disciplines/{discipline_id}")
def delete_discipline(
    discipline_id: int,
    session: Session = Depends(get_session),
    admin: User = Depends(require_admin)
):
    discipline = session.get(Discipline, discipline_id)
    if not discipline:
        raise HTTPException(status_code=404, detail="Discipline not found")
    
    try:
        session.delete(discipline)
        session.commit()
        return {"ok": True, "message": "Discipline deleted successfully"}
    except IntegrityError:
        session.rollback()
        # Fallback to soft delete
        discipline.active = False
        session.add(discipline)
        session.commit()
        return {"ok": True, "message": "Discipline deactivated (cannot be deleted due to existing student choices)"}

@router.post("/choices/clear")
def clear_all_choices(
    session: Session = Depends(get_session),
    admin: User = Depends(require_admin)
):
    try:
        session.exec(SQLModel.metadata.tables["choice"].delete())
        session.exec(SQLModel.metadata.tables["choiceset"].delete())
        session.commit()
        return {"ok": True, "message": "All student choices have been cleared"}
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))

from fastapi import UploadFile, File
import pandas as pd
import io

@router.post("/disciplines/import")
async def import_disciplines(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    admin: User = Depends(require_admin)
):
    try:
        content = await file.read()
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content))
        else:
            # For Excel, we might want to extract hyperlinks
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), data_only=False)
            ws = wb.active
            df = pd.read_excel(io.BytesIO(content))
        
        # mapping based on keywords to be more robust
        def get_col_name(cols, keywords):
            for col in cols:
                if not col or pd.isna(col): continue
                # Remove spaces and newlines for comparison
                clean_col = "".join(str(col).lower().split())
                if all("".join(k.lower().split()) in clean_col for k in keywords):
                    return col
            return None

        # Try to find the header row if the first one isn't it
        header_index = 0
        current_cols = list(df.columns)
        
        # Check if current headers are valid
        if not (get_col_name(current_cols, ["Код", "ВК"]) and get_col_name(current_cols, ["Назва", "дисциплін"])):
            # If not, scan first 10 rows to find headers
            found = False
            # Read without header to scan rows accurately
            df_source = pd.read_excel(io.BytesIO(content), header=None) if not file.filename.endswith('.csv') else pd.read_csv(io.BytesIO(content), header=None)
            
            for i in range(min(10, len(df_source))):
                row_values = df_source.iloc[i].tolist()
                if get_col_name(row_values, ["Код", "ВК"]) and get_col_name(row_values, ["Назва", "дисциплін"]):
                    # Found it! Header is row 'i', data starts at 'i+1'
                    header_index = i
                    # Use this source and set columns manually
                    headers = [str(col).strip() if not pd.isna(col) else f"Unnamed_{idx}" for idx, col in enumerate(row_values)]
                    df = df_source.iloc[i+1:].reset_index(drop=True)
                    df.columns = headers
                    if not file.filename.endswith('.csv'):
                        # Already loaded workbook above as 'wb', we'll use it
                        pass 
                    found = True
                    break
            
            if not found:
                raise HTTPException(status_code=400, detail=f"Не вдалося знайти рядок із заголовками (Код та Назва). Знайдено лише: {list(df.columns[:3])}...")
        else:
            # First row was actually headers, but we might still need openpyxl ws
            pass

        # Detect columns from final headers
        final_cols = list(df.columns)
        code_col = get_col_name(final_cols, ["Код", "ВК"]) or get_col_name(final_cols, ["Code"])
        title_col = get_col_name(final_cols, ["Назва", "дисциплін"]) or get_col_name(final_cols, ["Title"])
        credits_col = get_col_name(final_cols, ["кредит"]) or get_col_name(final_cols, ["Credits"])
        teacher_col = get_col_name(final_cols, ["викладач"]) or get_col_name(final_cols, ["Teacher"])
        comp_col = get_col_name(final_cols, ["компетентност"]) or get_col_name(final_cols, ["Competence"])
        comm_col = get_col_name(final_cols, ["коміс"]) or get_col_name(final_cols, ["Commission"])
        spec_col = get_col_name(final_cols, ["коду", "спеціальн"]) or get_col_name(final_cols, ["Specialty"])
        
        # URL column search: any of these keywords (aggressive search)
        url_col = None
        url_col_idx = -1
        for idx, col in enumerate(final_cols):
            c = str(col).lower()
            if any(k in c for k in ["диск", "disk", "google", "посилання", "силабус", "папк", "матеріал", "докум"]):
                url_col = col
                url_col_idx = idx
                break

        print(f"IMPORT DEBUG: Final Map: code={code_col}, title={title_col}, teacher={teacher_col}, credits={credits_col}, url={url_col}, comm={comm_col}, spec={spec_col}")

        def normalize_code(c):
            if not c or pd.isna(c): return ""
            # Replace common Latin lookalikes with Cyrillic and normalize dashes
            s = str(c).strip().upper()
            s = s.replace("V", "В").replace("B", "В").replace("K", "К").replace("–", "-").replace("—", "-")
            return s

        stats = {"created": 0, "updated": 0}
        
        for index, row in df.iterrows():
            raw_code = row.get(code_col)
            if pd.isna(raw_code) or pd.isna(row.get(title_col)):
                continue
                
            code = normalize_code(raw_code)
            # Skip if it's just the header or empty
            if not code or code in ["КОД", "КОДОСВІТНЬОГОКОМПОНЕНТА"]:
                continue

            # Normalized search in DB
            all_disciplines = session.exec(select(Discipline)).all()
            discipline = next((d for d in all_disciplines if normalize_code(d.code) == code), None)

            # TRY TO EXTRACT HYPERLINK IF EXCEL
            extracted_url = None
            if not file.filename.endswith('.csv') and url_col_idx != -1:
                try:
                    # openpyxl uses 1-based indexing
                    # header_index is where headers are. rows below are data.
                    # row index starts from 0 in iterrows, so it's data row.
                    # Excel row = header_index + 1 (for header) + index + 1
                    excel_row = header_index + 2 + index 
                    cell = ws.cell(row=excel_row, column=url_col_idx + 1)
                    if cell.hyperlink:
                        extracted_url = cell.hyperlink.target
                except:
                    pass

            doc_url = extracted_url or (str(row[url_col]).strip() if url_col and not pd.isna(row.get(url_col)) else None)
            
            data = {
                "title": str(row[title_col]).strip(),
                "short_info": str(row.get('short_info', '')) if not pd.isna(row.get('short_info')) else None,
                "doc_url": doc_url,
                "commission_name": str(row[comm_col]).strip() if comm_col and not pd.isna(row.get(comm_col)) else None,
                "spec": str(row[spec_col]).strip() if spec_col and not pd.isna(row.get(spec_col)) else None,
                "credits": float(row[credits_col]) if credits_col and not pd.isna(row.get(credits_col)) else None,
                "teacher": str(row[teacher_col]).strip() if teacher_col and not pd.isna(row.get(teacher_col)) else None,
                "ctype": str(row[comp_col]).strip() if comp_col and not pd.isna(row.get(comp_col)) else None,
            }
            
            if discipline:
                discipline.title = data["title"]
                discipline.doc_url = data["doc_url"]
                discipline.commission_name = data["commission_name"]
                discipline.specialty_code = data["spec"]
                discipline.credits = data["credits"]
                discipline.teacher_name = data["teacher"]
                discipline.competence_type = data["ctype"]
                discipline.updated_at = datetime.utcnow()
                stats["updated"] += 1
            else:
                new_disc = Discipline(
                    code=code, 
                    title=data["title"],
                    doc_url=data["doc_url"],
                    commission_name=data["commission_name"],
                    specialty_code=data["spec"],
                    credits=data["credits"],
                    teacher_name=data["teacher"],
                    competence_type=data["ctype"]
                )
                session.add(new_disc)
                stats["created"] += 1
                
        session.commit()
        return {"ok": True, "message": f"Імпорт завершено: створено {stats['created']}, оновлено {stats['updated']}"}
        
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"Помилка імпорту: {str(e)}")

